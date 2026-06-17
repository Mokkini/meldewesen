"""
Importiert die KW 24 (08.06–12.06.2026) Daten aus der Excel-Datei in die SQLite-DB.
Aufruf: python3 scripts/seed.py
"""
import openpyxl
import sqlite3
import os
from datetime import datetime, date

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'Meldewesen.xlsx')
DB_PATH    = os.path.join(os.path.dirname(__file__), '..', 'meldewesen.db')

WEEK_START = date(2026, 6, 8)
WEEK_END   = date(2026, 6, 12)

DEPOT_MAP = {
    'UP 22 LMO Delmenhorst':   'UP 22 – Delmenhorst',
    'UP 32 LMO Langenfeld':    'UP 32 – Langenfeld',
    'UP 34 LMO Melle':         'UP 34 – Melle',
    'UP 62 LMO Renningen':     'UP 62 – Renningen',
    'UP 68 LMO Aschaffenburg': 'UP 68 – Aschaffenburg',
    'UP 86 LMO Föhren':        'UP 86 – Föhren',
    'UP 90 LMO Güstrow':       'UP 90 – Güstrow',
    'UP 93 LMO Döbeln':        'UP 93 – Döbeln',
}

STATUS_MAP = {
    'zugestellt':           'zugestellt',
    'nicht zugestellt':     'nicht zugestellt',
    'nicht zugestellt ':    'nicht zugestellt',
    'Nicht zugestellt':     'nicht zugestellt',
    'nicht abgeholt':       'nicht abgeholt',
    'nicht geladen':        'nicht geladen',
    'verspätet':            'verspätet',
    'verspätete zustellung':'verspätet',
    'erneut einplanen':     'nicht zugestellt',
    'neu einplanen':        'nicht zugestellt',
    '-':                    'Tour-Umstellung',
    '':                     None,
}

def clean_status(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    return STATUS_MAP.get(s, s[:50] if s else None)

def clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s != 'None' else None

def clean_grund(grund, freitext=None):
    g = clean_str(grund)
    ft = clean_str(freitext)
    # Map common grounds to canonical categories
    if g:
        low = g.lower()
        if 'waz' in low:
            return 'WAZ verpasst', g
        if 'zeitmangel' in low:
            return 'Zeitmangel', g
        if 'geschlossen' in low or 'laden zu' in low:
            return 'Laden geschlossen', g
        if 'krank' in low or 'fahrerausfall' in low:
            return 'Fahrerausfall / Krankmeldung', g
        if 'dispofehl' in low:
            return 'Dispofehler', g
        if 'scanner' in low or 'manuell' in low:
            return 'Scannerprobleme / manuell geschlossen', g
        if 'adresse' in low or 'unkorrekt' in low:
            return 'Adresse unkorrekt', g
        if 'gesperrt' in low or 'markt' in low or 'fest' in low:
            return 'Straße gesperrt', g
        if 'vergessen' in low or 'fehlt' in low:
            return 'Ware fehlt / vergessen', g
        if 'tour' in low and ('von' in low or 'auf' in low):
            return 'Tour-Umstellung', g
    return 'Sonstiges', g or ft

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    lmo_sheets = [s for s in wb.sheetnames if 'LMO' in s]

    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    cur.execute('''
        CREATE TABLE IF NOT EXISTS meldungen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            datum TEXT NOT NULL,
            depot TEXT NOT NULL,
            tour TEXT,
            fahrer TEXT,
            unternehmer TEXT,
            typ TEXT NOT NULL DEFAULT "zusteller",
            sendungsnummer TEXT,
            grund TEXT,
            grund_freitext TEXT,
            erneute_zustellung TEXT,
            meldung_abrechnung TEXT DEFAULT "nein",
            tour_abgemeldet TEXT DEFAULT "ja",
            kunde TEXT,
            adresse TEXT,
            zustell_status TEXT,
            erstellt_am TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            erstellt_von TEXT
        )
    ''')

    inserted = 0
    skipped  = 0

    for sheet_name in lmo_sheets:
        ws    = wb[sheet_name]
        depot = DEPOT_MAP.get(sheet_name, sheet_name)

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # skip header

            datum_raw = row[0]
            if not isinstance(datum_raw, datetime):
                skipped += 1
                continue

            d = datum_raw.date()
            if not (WEEK_START <= d <= WEEK_END):
                skipped += 1
                continue

            # Determine zusteller vs abholer
            zusteller = row[4]
            abholer   = row[5]
            typ = 'abholer' if abholer and str(abholer).strip().lower() == 'x' else 'zusteller'

            # Sendungsnummer
            sdg = clean_str(row[6])

            # Status
            status_raw = clean_str(row[13])
            zustell_status = clean_status(status_raw)

            # Grund
            grund_kat, grund_ft = clean_grund(row[7])

            # Erneute Zustellung
            erneute = clean_str(row[8])
            if erneute in ('ja', 'Ja', 'Ja ', 'x', 'X'):
                erneute = 'ja'
            elif erneute in ('nein', 'Nein'):
                erneute = 'nein'
            else:
                erneute = 'nein'

            # Meldung Abrechnung
            abrechn = clean_str(row[9])
            meldung_abrechnung = 'ja' if abrechn and abrechn.lower() == 'ja' else 'nein'

            # Tour abgemeldet
            tab = clean_str(row[10])
            tour_abgemeldet = 'ja' if tab and tab.lower() == 'ja' else 'nein'

            # Adresse (handle #VALUE! from Excel formula error)
            adresse_raw = row[12]
            adresse = None if str(adresse_raw).startswith('#') else clean_str(adresse_raw)

            cur.execute('''
                INSERT INTO meldungen
                    (datum, depot, tour, fahrer, unternehmer, typ, sendungsnummer,
                     grund, grund_freitext, erneute_zustellung, meldung_abrechnung,
                     tour_abgemeldet, kunde, adresse, zustell_status, erstellt_von)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                d.isoformat(),
                depot,
                clean_str(row[1]),
                clean_str(row[2]),
                clean_str(row[3]),
                typ,
                sdg,
                grund_kat,
                grund_ft,
                erneute,
                meldung_abrechnung,
                tour_abgemeldet,
                clean_str(row[11]),
                adresse,
                zustell_status,
                'Import KW24',
            ))
            inserted += 1

    conn.commit()
    conn.close()
    print(f"✅ Import fertig: {inserted} Meldungen importiert, {skipped} Zeilen übersprungen.")

if __name__ == '__main__':
    main()
